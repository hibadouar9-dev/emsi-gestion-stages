from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import InscriptionForm, ConnexionForm , ConventionForm, OffreStageForm
from .models import Utilisateur, Convention, OffreStage, Candidature

def inscription(request):
    if request.method == 'POST':
        form = InscriptionForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'etudiant'
            user.save()
            login(request, user)
            messages.success(request, f'Bienvenue {user.username} ! Votre compte est créé.')
            return redirect('dashboard')
    else:
        form = InscriptionForm()
    return render(request, 'stages/inscription.html', {'form': form})

def connexion(request):
    if request.method == 'POST':
        form = ConnexionForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Bonjour {user.username} !')
                return redirect('dashboard')  # ← TOUS vers dashboard
        messages.error(request, 'Identifiants incorrects')
    else:
        form = ConnexionForm()
    return render(request, 'stages/connexion.html', {'form': form})

def deconnexion(request):
    logout(request)
    messages.info(request, 'Vous êtes déconnecté.')
    return redirect('connexion')

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Utilisateur, Convention, OffreStage, Evaluation

@login_required
def dashboard(request):
    user = request.user
    
    if user.role == 'etudiant':
        conventions = Convention.objects.filter(etudiant=user)
        candidatures = Candidature.objects.filter(etudiant=user).count()
        
        context = {
            'total_conventions': conventions.count(),
            'conventions_validees': conventions.filter(statut='validee').count(),
            'total_offres': OffreStage.objects.filter(active=True).count(),
            'conventions_attente': conventions.filter(statut='attente_service').count(),
            'dernieres_conventions': conventions.order_by('-date_depot')[:5],
            'candidatures': candidatures,
            'user': user,
        }
        return render(request, 'stages/dashboard.html', context)
    
    elif user.role == 'service':
        context = {
            'total_conventions': Convention.objects.count(),
            'conventions_attente': Convention.objects.filter(statut='attente_service').count(),
            'conventions_validees': Convention.objects.filter(statut='validee').count(),
            'offres_actives': OffreStage.objects.filter(active=True).count(),
            'dernieres_conventions': Convention.objects.order_by('-date_depot')[:5],
            'user': user,
        }
        return render(request, 'stages/dashboard_service.html', context)
    
    else:
        context = {
            'total_conventions': Convention.objects.count(),
            'conventions_validees': Convention.objects.filter(statut='validee').count(),
            'total_offres': OffreStage.objects.filter(active=True).count(),
            'dernieres_conventions': Convention.objects.order_by('-date_depot')[:5],
            'user': user,
        }
        return render(request, 'stages/dashboard.html', context)

@login_required
def depot_convention(request):
    if request.user.role != 'etudiant':
        messages.error(request, 'Seuls les étudiants peuvent déposer une convention.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = ConventionForm(request.POST, request.FILES)
        if form.is_valid():
            convention = form.save(commit=False)
            convention.etudiant = request.user
            convention.save()
            messages.success(request, '✅ Convention déposée avec succès ! En attente de validation.')
            return redirect('mes_conventions')
        else:
            messages.error(request, '❌ Erreur dans le formulaire. Vérifiez les champs.')
    else:
        form = ConventionForm()
    
    return render(request, 'stages/depot_convention.html', {'form': form})

@login_required
def mes_conventions(request):
    if request.user.role != 'etudiant':
        return redirect('dashboard')
    
    conventions = Convention.objects.filter(etudiant=request.user).order_by('-date_depot')
    return render(request, 'stages/mes_conventions.html', {'conventions': conventions})

from datetime import date  # Ajoutez cet import en haut du fichier s'il n'existe pas



@login_required
def conventions_a_valider(request):
    user = request.user
    if user.role == 'tuteur':
        # Montrer les conventions en attente du tuteur ainsi que les conventions déjà validées
        conventions = Convention.objects.filter(statut__in=['attente_tuteur', 'validee'])
    elif user.role == 'responsable':
        conventions = Convention.objects.filter(statut='attente_responsable')
    elif user.role == 'service':
        conventions = Convention.objects.filter(statut='attente_service')
    else:
        messages.error(request, 'Vous n\'avez pas accès à cette page')
        return redirect('dashboard')
    
    # Précharger les évaluations liées pour éviter les erreurs côté template
    evaluations = Evaluation.objects.filter(convention__in=conventions)
    eval_map = {e.convention_id: e for e in evaluations}
    # Attacher un attribut 'evaluation' à chaque convention (None si pas d'évaluation)
    conventions = list(conventions)  # évaluer la queryset
    for conv in conventions:
        conv.evaluation = eval_map.get(conv.id)

    # AJOUTEZ 'now' DANS LE CONTEXTE
    return render(request, 'stages/conventions_a_valider.html', {
        'conventions': conventions,
        'now': date.today(),
    })

@login_required
def valider_convention(request, convention_id):
    convention = get_object_or_404(Convention, id=convention_id)
    user = request.user
    
    if user.role == 'tuteur' and convention.statut == 'attente_tuteur':
        convention.statut = 'attente_responsable'
        convention.save()
        messages.success(request, f'✅ Convention validée. Transmise au responsable.')
    elif user.role == 'responsable' and convention.statut == 'attente_responsable':
        convention.statut = 'attente_service'
        convention.save()
        messages.success(request, f'✅ Convention validée. Transmise au service des stages.')
    elif user.role == 'service' and convention.statut == 'attente_service':
        convention.statut = 'validee'
        convention.save()
        messages.success(request, f'✅ Convention validée définitivement !')
    else:
        messages.error(request, '❌ Vous ne pouvez pas valider cette convention')
    
    return redirect('conventions_a_valider')

@login_required
def refuser_convention(request, convention_id):
    convention = get_object_or_404(Convention, id=convention_id)
    
    if request.method == 'POST':
        commentaire = request.POST.get('commentaire', '')
        convention.statut = 'refusee'
        convention.commentaire_refus = commentaire
        convention.save()
        messages.warning(request, f'❌ Convention refusée : {commentaire}')
        return redirect('conventions_a_valider')
    
    return render(request, 'stages/refuser_convention.html', {'convention': convention})

from .models import OffreStage
from .forms import OffreStageForm

# === GESTION DES OFFRES (SERVICE STAGES) ===
@login_required
def liste_offres_service(request):
    if request.user.role != 'service':
        messages.error(request, 'Accès non autorisé')
        return redirect('dashboard')
    
    offres = OffreStage.objects.all().order_by('-date_creation')
    return render(request, 'stages/service_offres.html', {'offres': offres})

@login_required
def creer_offre(request):
    if request.user.role != 'service':
        messages.error(request, 'Accès non autorisé')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = OffreStageForm(request.POST)
        if form.is_valid():
            offre = form.save(commit=False)
            offre.entreprise = request.user
            offre.save()
            messages.success(request, f'✅ Offre "{offre.titre}" créée avec succès !')
            return redirect('liste_offres_service')
        else:
            messages.error(request, 'Erreur dans le formulaire')
    else:
        form = OffreStageForm()
    
    return render(request, 'stages/creer_offre.html', {'form': form})

@login_required
def modifier_offre(request, offre_id):
    if request.user.role != 'service':
        messages.error(request, 'Accès non autorisé')
        return redirect('dashboard')
    
    offre = get_object_or_404(OffreStage, id=offre_id)
    
    if request.method == 'POST':
        form = OffreStageForm(request.POST, instance=offre)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Offre "{offre.titre}" modifiée avec succès !')
            return redirect('liste_offres_service')
    else:
        form = OffreStageForm(instance=offre)
    
    return render(request, 'stages/modifier_offre.html', {'form': form, 'offre': offre})

@login_required
def supprimer_offre(request, offre_id):
    if request.user.role != 'service':
        messages.error(request, 'Accès non autorisé')
        return redirect('dashboard')
    
    offre = get_object_or_404(OffreStage, id=offre_id)
    
    if request.method == 'POST':
        titre = offre.titre
        offre.delete()
        messages.success(request, f'✅ Offre "{titre}" supprimée avec succès !')
        return redirect('liste_offres_service')
    
    return render(request, 'stages/supprimer_offre.html', {'offre': offre})

# === CONSULTATION DES OFFRES (ETUDIANTS) ===
@login_required
def consulter_offres(request):
    if request.user.role != 'etudiant':
        messages.error(request, 'Accès non autorisé')
        return redirect('dashboard')
    
    offres = OffreStage.objects.filter(active=True).order_by('-date_creation')
    return render(request, 'stages/consulter_offres.html', {'offres': offres})

from .models import Candidature

@login_required
def postuler_offre(request, offre_id):
    if request.user.role != 'etudiant':
        messages.error(request, 'Seuls les étudiants peuvent postuler')
        return redirect('dashboard')
    
    offre = get_object_or_404(OffreStage, id=offre_id, active=True)
    
    # Vérifier si l'étudiant a déjà postulé
    deja_postule = Candidature.objects.filter(etudiant=request.user, offre=offre).exists()
    
    if request.method == 'POST':
        if deja_postule:
            messages.warning(request, 'Vous avez déjà postulé à cette offre')
        else:
            message = request.POST.get('message', '')
            Candidature.objects.create(
                etudiant=request.user,
                offre=offre,
                message=message
            )
            messages.success(request, f'✅ Candidature envoyée pour "{offre.titre}"')
        return redirect('consulter_offres')
    
    return render(request, 'stages/postuler_offre.html', {'offre': offre, 'deja_postule': deja_postule})

@login_required
def mes_candidatures(request):
    if request.user.role != 'etudiant':
        return redirect('dashboard')
    
    candidatures = Candidature.objects.filter(etudiant=request.user).order_by('-date_candidature')
    return render(request, 'stages/mes_candidatures.html', {'candidatures': candidatures})

@login_required
def gerer_candidatures(request):
    if request.user.role != 'service':
        messages.error(request, 'Accès non autorisé')
        return redirect('dashboard')
    
    candidatures = Candidature.objects.all().order_by('-date_candidature')
    return render(request, 'stages/gerer_candidatures.html', {'candidatures': candidatures})

@login_required
def modifier_statut_candidature(request, candidature_id, statut):
    if request.user.role != 'service':
        messages.error(request, 'Accès non autorisé')
        return redirect('dashboard')
    
    candidature = get_object_or_404(Candidature, id=candidature_id)
    if statut in ['acceptee', 'refusee']:
        candidature.statut = statut
        candidature.save()
        messages.success(request, f'Candidature {statut} pour {candidature.etudiant.username}')
    
    return redirect('gerer_candidatures')

def accueil(request):
    return render(request, 'stages/accueil.html')

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

@login_required
def export_excel(request):
    if request.user.role != 'service':
        messages.error(request, 'Accès non autorisé')
        return redirect('dashboard')
    
    # Récupérer les conventions validées
    conventions = Convention.objects.filter(statut='validee').order_by('-date_depot')
    
    # Créer le classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conventions validées"
    
    # En-têtes
    headers = ['Étudiant', 'Email', 'Filière', 'Entreprise', 'Tuteur entreprise', 
               'Date début', 'Date fin', 'Date dépôt', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Données
    for row, conv in enumerate(conventions, 2):
        ws.cell(row=row, column=1, value=conv.etudiant.username)
        ws.cell(row=row, column=2, value=conv.etudiant.email)
        ws.cell(row=row, column=3, value=conv.etudiant.filiere or '-')
        ws.cell(row=row, column=4, value=conv.entreprise_nom)
        ws.cell(row=row, column=5, value=conv.tuteur_entreprise_nom)
        ws.cell(row=row, column=6, value=conv.date_debut.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=7, value=conv.date_fin.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=8, value=conv.date_depot.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=9, value=conv.get_statut_display())
    
    # Ajuster les colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=conventions_validees.xlsx'
    wb.save(response)
    
    return response

from datetime import date
from .models import Evaluation

from datetime import date

@login_required
def evaluer_convention(request, convention_id):
    convention = get_object_or_404(Convention, id=convention_id)
    
    # Vérifier que c'est un tuteur
    if request.user.role != 'tuteur':
        messages.error(request, 'Seuls les tuteurs peuvent évaluer')
        return redirect('dashboard')
    
    # Vérifier que la convention est validée
    if convention.statut != 'validee':
        messages.error(request, 'Cette convention n\'est pas encore validée')
        return redirect('dashboard')
    
    # Vérifier que le stage est terminé
    if convention.date_fin > date.today():
        messages.error(request, f'L\'évaluation sera disponible après le {convention.date_fin.strftime("%d/%m/%Y")}')
        return redirect('dashboard')
    
    # Récupérer ou créer l'évaluation
    evaluation, created = Evaluation.objects.get_or_create(convention=convention)
    
    if request.method == 'POST':
        evaluation.note_tuteur_emsi = request.POST.get('note')
        evaluation.commentaire_tuteur_emsi = request.POST.get('commentaire')
        evaluation.save()
        messages.success(request, f'Évaluation enregistrée pour {convention.etudiant.username}')
        return redirect('dashboard')
    
    return render(request, 'stages/evaluation.html', {
        'convention': convention,
        'evaluation': evaluation,
    })